"""
Celery tasks for the EnterpriseLand platform.

Provides asynchronous task processing for audit logging,
notifications, and other background operations.
"""

import logging
from typing import Optional, Dict, Any
from celery import shared_task
from django.contrib.auth import get_user_model
from django.apps import apps

from accounts.models import Group
from core.models import AuditLog

User = get_user_model()
logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def create_audit_log_async(self, **kwargs):
    """
    Create an audit log entry asynchronously.
    
    Args:
        **kwargs: Arguments for AuditLog.objects.create_log()
        
    Returns:
        ID of created audit log or None if failed
    """
    try:
        # Handle user and group references that need to be resolved
        user_id = kwargs.pop('user_id', None)
        user = None
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                kwargs['user'] = user
            except User.DoesNotExist:
                logger.warning(f"User {user_id} not found for audit log")
        
        group_id = kwargs.pop('group_id', None)
        group = None
        if group_id:
            try:
                group = Group.objects.get(id=group_id)
                kwargs['group'] = group
            except Group.DoesNotExist:
                logger.warning(f"Group {group_id} not found for audit log")
        
        # Handle content_object if provided
        content_type_id = kwargs.pop('content_type_id', None)
        object_id = kwargs.pop('object_id', None)
        content_object = None
        
        if content_type_id and object_id:
            try:
                from django.contrib.contenttypes.models import ContentType
                content_type = ContentType.objects.get(id=content_type_id)
                model_class = content_type.model_class()
                content_object = model_class.objects.get(pk=object_id)
                kwargs['content_object'] = content_object
            except Exception as e:
                logger.warning(f"Content object not found: {e}")
        
        # Create the audit log
        audit_log = AuditLog.objects.create_log(**kwargs)
        
        if audit_log:
            logger.debug(f"Created audit log {audit_log.id} asynchronously")
            return str(audit_log.id)
        else:
            logger.error("Failed to create audit log")
            return None
            
    except Exception as exc:
        logger.error(f"Error creating audit log: {exc}")
        
        # Retry the task
        if self.request.retries < self.max_retries:
            logger.info(f"Retrying audit log creation (attempt {self.request.retries + 1})")
            raise self.retry(countdown=60, exc=exc)
        else:
            logger.error(f"Max retries exceeded for audit log creation: {exc}")
            return None


@shared_task
def cleanup_old_audit_logs(days_to_keep: int = 90):
    """
    Clean up old audit logs based on retention policy.
    
    Args:
        days_to_keep: Number of days to keep audit logs
        
    Returns:
        Number of deleted audit logs
    """
    try:
        from django.utils import timezone
        from datetime import timedelta
        
        cutoff_date = timezone.now() - timedelta(days=days_to_keep)
        
        # Delete old audit logs
        deleted_count, _ = AuditLog.objects.filter(
            timestamp__lt=cutoff_date
        ).delete()
        
        logger.info(f"Cleaned up {deleted_count} audit logs older than {days_to_keep} days")
        return deleted_count
        
    except Exception as e:
        logger.error(f"Error cleaning up audit logs: {e}")
        return 0


@shared_task
def generate_audit_report(
    report_type: str,
    start_date: str,
    end_date: str,
    user_id: Optional[str] = None,
    group_id: Optional[str] = None,
    export_format: str = 'json'
):
    """
    Generate an audit report asynchronously.
    
    Args:
        report_type: Type of report to generate
        start_date: Start date for the report (ISO format)
        end_date: End date for the report (ISO format)
        user_id: Optional user ID to filter by
        group_id: Optional group ID to filter by
        export_format: Export format (json, csv, excel)
        
    Returns:
        Report data or file path
    """
    try:
        from datetime import datetime
        from django.core.serializers.json import DjangoJSONEncoder
        import json
        
        # Parse dates
        start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        # Build query
        queryset = AuditLog.objects.filter(
            timestamp__gte=start_dt,
            timestamp__lte=end_dt
        )
        
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        if group_id:
            queryset = queryset.filter(group_id=group_id)
        
        # Apply report type filters
        if report_type == 'security':
            queryset = queryset.filter(
                action__in=[
                    AuditLog.Action.LOGIN_FAILED,
                    AuditLog.Action.PERMISSION_CHANGE,
                    AuditLog.Action.PASSWORD_CHANGE,
                    AuditLog.Action.DELETE,
                    AuditLog.Action.ADMIN_ACCESS
                ]
            )
        elif report_type == 'errors':
            queryset = queryset.filter(success=False)
        elif report_type == 'data_changes':
            queryset = queryset.filter(
                action__in=[
                    AuditLog.Action.CREATE,
                    AuditLog.Action.UPDATE,
                    AuditLog.Action.DELETE
                ]
            ).exclude(changes={})
        
        # Get the data
        audit_logs = queryset.order_by('-timestamp')[:10000]  # Limit to 10k records
        
        # Format the data
        report_data = []
        for log in audit_logs:
            report_data.append({
                'id': str(log.id),
                'timestamp': log.timestamp.isoformat(),
                'action': log.action,
                'user': log.user.email if log.user else None,
                'model_name': log.model_name,
                'object_id': log.object_id,
                'ip_address': log.ip_address,
                'success': log.success,
                'changes': log.mask_sensitive_data(),
                'error_message': log.error_message,
                'metadata': log.metadata
            })
        
        if export_format == 'json':
            return json.dumps(report_data, cls=DjangoJSONEncoder, indent=2)
        elif export_format == 'csv':
            return _export_to_csv(report_data)
        elif export_format == 'excel':
            return _export_to_excel(report_data)
        else:
            return report_data
            
    except Exception as e:
        logger.error(f"Error generating audit report: {e}")
        return None


def _export_to_csv(data: list) -> str:
    """Export audit data to CSV format."""
    try:
        import csv
        import io
        
        output = io.StringIO()
        if not data:
            return ""
        
        fieldnames = data[0].keys()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in data:
            # Flatten complex fields
            flattened_row = {}
            for key, value in row.items():
                if isinstance(value, (dict, list)):
                    flattened_row[key] = json.dumps(value)
                else:
                    flattened_row[key] = value
            writer.writerow(flattened_row)
        
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error exporting to CSV: {e}")
        return ""


def _export_to_excel(data: list) -> bytes:
    """Export audit data to Excel format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Log Report"
        
        if not data:
            return b""
        
        # Headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data[header]
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        return excel_buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        return b""


@shared_task
def archive_old_audit_logs(days_to_archive: int = 365):
    """
    Archive old audit logs to cold storage.
    
    Args:
        days_to_archive: Number of days after which to archive logs
        
    Returns:
        Number of archived audit logs
    """
    try:
        from django.utils import timezone
        from datetime import timedelta
        import json
        
        cutoff_date = timezone.now() - timedelta(days=days_to_archive)
        
        # Get old audit logs
        old_logs = AuditLog.objects.filter(
            timestamp__lt=cutoff_date
        ).select_related('user', 'group', 'content_type')
        
        if not old_logs.exists():
            logger.info("No audit logs to archive")
            return 0
        
        # Archive to JSON files (could be S3, etc.)
        archive_data = []
        for log in old_logs:
            archive_data.append({
                'id': str(log.id),
                'timestamp': log.timestamp.isoformat(),
                'action': log.action,
                'user_email': log.user.email if log.user else None,
                'model_name': log.model_name,
                'object_id': log.object_id,
                'changes': log.changes,
                'ip_address': log.ip_address,
                'user_agent': log.user_agent,
                'group_name': log.group.name if log.group else None,
                'metadata': log.metadata,
                'success': log.success,
                'error_message': log.error_message
            })
        
        # Save archive (this could be enhanced to save to S3, etc.)
        archive_filename = f"audit_archive_{cutoff_date.strftime('%Y%m%d')}.json"
        # In a real implementation, you'd save this to your archive storage
        
        count = old_logs.count()
        logger.info(f"Would archive {count} audit logs to {archive_filename}")
        
        # For now, just log that we would archive (don't actually delete)
        # In production, you'd want to:
        # 1. Save archive_data to S3/cold storage
        # 2. Verify the archive was saved successfully
        # 3. Then delete the old records: old_logs.delete()
        
        return count
        
    except Exception as e:
        logger.error(f"Error archiving audit logs: {e}")
        return 0


@shared_task
def audit_log_health_check():
    """
    Perform health check on audit logging system.
    
    Returns:
        Health check results
    """
    try:
        from django.utils import timezone
        from datetime import timedelta
        
        now = timezone.now()
        results = {
            'timestamp': now.isoformat(),
            'status': 'healthy',
            'checks': {}
        }
        
        # Check recent audit log creation
        recent_logs = AuditLog.objects.filter(
            timestamp__gte=now - timedelta(hours=1)
        ).count()
        
        results['checks']['recent_logs'] = {
            'count': recent_logs,
            'status': 'healthy' if recent_logs > 0 else 'warning',
            'message': f"Found {recent_logs} audit logs in the last hour"
        }
        
        # Check database connectivity
        try:
            total_logs = AuditLog.objects.count()
            results['checks']['database'] = {
                'status': 'healthy',
                'total_logs': total_logs,
                'message': f"Database accessible, {total_logs} total audit logs"
            }
        except Exception as e:
            results['checks']['database'] = {
                'status': 'error',
                'message': f"Database error: {e}"
            }
            results['status'] = 'unhealthy'
        
        # Check for failed audit logs
        failed_logs = AuditLog.objects.filter(
            timestamp__gte=now - timedelta(hours=24),
            success=False
        ).count()
        
        results['checks']['failed_operations'] = {
            'count': failed_logs,
            'status': 'healthy' if failed_logs < 100 else 'warning',
            'message': f"Found {failed_logs} failed operations in the last 24 hours"
        }
        
        # Check audit log entries growth
        logs_today = AuditLog.objects.filter(
            timestamp__gte=now.replace(hour=0, minute=0, second=0, microsecond=0)
        ).count()
        
        results['checks']['growth_rate'] = {
            'logs_today': logs_today,
            'status': 'healthy' if logs_today < 10000 else 'warning',
            'message': f"Created {logs_today} audit logs today"
        }
        
        logger.info(f"Audit log health check completed: {results['status']}")
        return results
        
    except Exception as e:
        logger.error(f"Error in audit log health check: {e}")
        return {
            'timestamp': timezone.now().isoformat(),
            'status': 'error',
            'message': str(e)
        }